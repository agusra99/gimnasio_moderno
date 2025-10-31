'''"""
views/reports.py
Vista para generar reportes y estadísticas del gimnasio.
Incluye gráficos, exportación a PDF y Excel.
"""

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QLabel, QHBoxLayout, QPushButton,
    QComboBox, QGroupBox, QTableWidget, QTableWidgetItem, QMessageBox
)
from PySide6.QtCore import Qt
import sqlite3
from datetime import datetime, timedelta
from collections import defaultdict


class ReportsView(QWidget):
    def __init__(self, db_path):
        super().__init__()
        self.db_path = db_path
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()

        # Título
        title = QLabel("📊 Reportes y Estadísticas")
        title.setStyleSheet("font-size: 22px; font-weight: bold; margin: 10px;")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        # Panel de filtros
        filtros_group = QGroupBox("🔍 Seleccionar Período")
        filtros_layout = QHBoxLayout()

        self.combo_periodo = QComboBox()
        self.combo_periodo.addItems([
            "Mes Actual",
            "Mes Anterior",
            "Últimos 3 Meses",
            "Últimos 6 Meses",
            "Año Actual",
            "Año Anterior",
            "Últimos 12 Meses",
            "Todo el Tiempo"
        ])
        self.combo_periodo.currentIndexChanged.connect(self.generar_reporte)

        btn_generar = QPushButton("📈 Generar Reporte")
        btn_generar.setStyleSheet("""
            QPushButton {
                padding: 10px 20px;
                background: #2196F3;
                color: white;
                border-radius: 6px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #1976D2;
            }
        """)
        btn_generar.clicked.connect(self.generar_reporte)

        filtros_layout.addWidget(QLabel("Período:"))
        filtros_layout.addWidget(self.combo_periodo)
        filtros_layout.addWidget(btn_generar)
        filtros_layout.addStretch()
        filtros_group.setLayout(filtros_layout)
        layout.addWidget(filtros_group)

        # Panel de resumen financiero
        resumen_group = QGroupBox("💰 Resumen Financiero")
        resumen_layout = QHBoxLayout()

        self.card_ingresos = self._crear_card("Total Ingresos", "$0", "#4CAF50")
        self.card_socios_pagaron = self._crear_card("Socios que Pagaron", "0", "#2196F3")
        self.card_promedio = self._crear_card("Promedio por Pago", "$0", "#FF9800")

        resumen_layout.addWidget(self.card_ingresos)
        resumen_layout.addWidget(self.card_socios_pagaron)
        resumen_layout.addWidget(self.card_promedio)
        resumen_group.setLayout(resumen_layout)
        layout.addWidget(resumen_group)

        # Panel de estadísticas de socios
        socios_group = QGroupBox("👥 Estadísticas de Socios")
        socios_layout = QHBoxLayout()

        self.card_total_socios = self._crear_card("Total Socios", "0", "#9C27B0")
        self.card_nuevos = self._crear_card("Nuevos en Período", "0", "#00BCD4")
        self.card_activos = self._crear_card("Activos", "0", "#4CAF50")

        socios_layout.addWidget(self.card_total_socios)
        socios_layout.addWidget(self.card_nuevos)
        socios_layout.addWidget(self.card_activos)
        socios_group.setLayout(socios_layout)
        layout.addWidget(socios_group)

        # Tabla de detalles
        detalles_group = QGroupBox("📋 Detalle de Pagos del Período")
        detalles_layout = QVBoxLayout()

        self.tabla = QTableWidget()
        self.tabla.setColumnCount(5)
        self.tabla.setHorizontalHeaderLabels([
            "Fecha", "Socio", "Monto", "Mes Correspondiente", "Plan"
        ])
        self.tabla.horizontalHeader().setStretchLastSection(True)
        self.tabla.setAlternatingRowColors(True)
        self.tabla.setStyleSheet("""
            QTableWidget {
                border: 1px solid #ddd;
                border-radius: 6px;
            }
            QHeaderView::section {
                background-color: #0078d7;
                color: white;
                padding: 8px;
                font-weight: bold;
            }
        """)

        detalles_layout.addWidget(self.tabla)
        detalles_group.setLayout(detalles_layout)
        layout.addWidget(detalles_group)

        # Botones de exportación
        export_layout = QHBoxLayout()

        self.btn_export_excel = QPushButton("📊 Exportar a Excel")
        self.btn_export_excel.setStyleSheet("""
            QPushButton {
                padding: 10px 20px;
                background: #4CAF50;
                color: white;
                border-radius: 6px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #388E3C;
            }
        """)
        self.btn_export_excel.clicked.connect(self.exportar_excel)

        self.btn_export_pdf = QPushButton("📄 Exportar a PDF")
        self.btn_export_pdf.setStyleSheet("""
            QPushButton {
                padding: 10px 20px;
                background: #F44336;
                color: white;
                border-radius: 6px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #D32F2F;
            }
        """)
        self.btn_export_pdf.clicked.connect(self.exportar_pdf)

        export_layout.addStretch()
        export_layout.addWidget(self.btn_export_excel)
        export_layout.addWidget(self.btn_export_pdf)
        layout.addLayout(export_layout)

        self.setLayout(layout)
        
        # Generar reporte inicial
        self.generar_reporte()

    def _crear_card(self, titulo, valor, color):
        """Crea una tarjeta de estadística."""
        from PySide6.QtWidgets import QFrame
        
        frame = QFrame()
        frame.setStyleSheet(f"""
            QFrame {{
                background: white;
                border-left: 5px solid {color};
                border-radius: 8px;
                padding: 20px;
            }}
        """)
        
        frame_layout = QVBoxLayout(frame)
        
        lbl_titulo = QLabel(titulo)
        lbl_titulo.setStyleSheet(f"font-size: 14px; color: #666; font-weight: 500;")
        
        lbl_valor = QLabel(valor)
        lbl_valor.setStyleSheet(f"font-size: 32px; font-weight: bold; color: {color};")
        lbl_valor.setAlignment(Qt.AlignCenter)
        
        frame_layout.addWidget(lbl_titulo)
        frame_layout.addWidget(lbl_valor)
        
        frame.lbl_valor = lbl_valor
        return frame

    def _obtener_rango_fechas(self):
        """Obtiene el rango de fechas según el período seleccionado."""
        hoy = datetime.now()
        periodo = self.combo_periodo.currentText()

        if periodo == "Mes Actual":
            inicio = hoy.replace(day=1).strftime('%Y-%m-%d')
            fin = hoy.strftime('%Y-%m-%d')
        elif periodo == "Mes Anterior":
            primer_dia_mes = hoy.replace(day=1)
            ultimo_mes = primer_dia_mes - timedelta(days=1)
            inicio = ultimo_mes.replace(day=1).strftime('%Y-%m-%d')
            fin = ultimo_mes.strftime('%Y-%m-%d')
        elif periodo == "Últimos 3 Meses":
            inicio = (hoy - timedelta(days=90)).strftime('%Y-%m-%d')
            fin = hoy.strftime('%Y-%m-%d')
        elif periodo == "Últimos 6 Meses":
            inicio = (hoy - timedelta(days=180)).strftime('%Y-%m-%d')
            fin = hoy.strftime('%Y-%m-%d')
        elif periodo == "Año Actual":
            inicio = hoy.replace(month=1, day=1).strftime('%Y-%m-%d')
            fin = hoy.strftime('%Y-%m-%d')
        elif periodo == "Año Anterior":
            inicio = hoy.replace(year=hoy.year-1, month=1, day=1).strftime('%Y-%m-%d')
            fin = hoy.replace(year=hoy.year-1, month=12, day=31).strftime('%Y-%m-%d')
        elif periodo == "Últimos 12 Meses":
            inicio = (hoy - timedelta(days=365)).strftime('%Y-%m-%d')
            fin = hoy.strftime('%Y-%m-%d')
        else:  # Todo el Tiempo
            inicio = "2000-01-01"
            fin = hoy.strftime('%Y-%m-%d')

        return inicio, fin

    def generar_reporte(self):
        """Genera el reporte con los datos del período seleccionado."""
        try:
            inicio, fin = self._obtener_rango_fechas()
            
            conn = sqlite3.connect(self.db_path)
            c = conn.cursor()

            # Ingresos totales del período
            c.execute("""
                SELECT SUM(monto), COUNT(*), COUNT(DISTINCT socio_id)
                FROM pagos
                WHERE fecha_pago BETWEEN ? AND ?
            """, (inicio, fin))
            
            resultado = c.fetchone()
            total_ingresos = resultado[0] if resultado[0] else 0
            cantidad_pagos = resultado[1]
            socios_pagaron = resultado[2]

            # Promedio por pago
            promedio = total_ingresos / cantidad_pagos if cantidad_pagos > 0 else 0

            # Actualizar cards financieros
            self.card_ingresos.lbl_valor.setText(f"${total_ingresos:,.2f}")
            self.card_socios_pagaron.lbl_valor.setText(str(socios_pagaron))
            self.card_promedio.lbl_valor.setText(f"${promedio:,.2f}")

            # Estadísticas de socios
            c.execute("SELECT COUNT(*) FROM socios")
            total_socios = c.fetchone()[0]

            # CORRECCIÓN: Usar fecha_inscripcion en lugar de fecha_alta
            c.execute("""
                SELECT COUNT(*) FROM socios
                WHERE fecha_inscripcion BETWEEN ? AND ?
            """, (inicio, fin))
            nuevos_socios = c.fetchone()[0]

            # Socios activos (con al menos un pago en el período)
            c.execute("""
                SELECT COUNT(DISTINCT socio_id) FROM pagos
                WHERE fecha_pago BETWEEN ? AND ?
            """, (inicio, fin))
            socios_activos = c.fetchone()[0]

            # Actualizar cards de socios
            self.card_total_socios.lbl_valor.setText(str(total_socios))
            self.card_nuevos.lbl_valor.setText(str(nuevos_socios))
            self.card_activos.lbl_valor.setText(str(socios_activos))

            # Cargar detalle de pagos
            c.execute("""
                SELECT p.fecha_pago, s.nombre || ' ' || s.apellido,
                       p.monto, p.mes_correspondiente, pl.nombre
                FROM pagos p
                JOIN socios s ON p.socio_id = s.id
                LEFT JOIN planes pl ON s.plan_id = pl.id
                WHERE p.fecha_pago BETWEEN ? AND ?
                ORDER BY p.fecha_pago DESC
            """, (inicio, fin))

            pagos = c.fetchall()
            self.tabla.setRowCount(len(pagos))

            for row, pago in enumerate(pagos):
                self.tabla.setItem(row, 0, QTableWidgetItem(pago[0]))
                self.tabla.setItem(row, 1, QTableWidgetItem(pago[1]))
                self.tabla.setItem(row, 2, QTableWidgetItem(f"${pago[2]:,.2f}"))
                self.tabla.setItem(row, 3, QTableWidgetItem(pago[3]))
                self.tabla.setItem(row, 4, QTableWidgetItem(pago[4] if pago[4] else "Sin plan"))

            conn.close()

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al generar reporte: {str(e)}")

    def exportar_excel(self):
        """Exporta el reporte a Excel con formato."""
        try:
            # Intentar usar openpyxl para Excel real
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                filename = f"reporte_gimnasio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                
                # Crear libro de Excel
                wb = Workbook()
                ws = wb.active
                ws.title = "Reporte de Pagos"
                
                # Escribir encabezados con formato
                headers = ["Fecha", "Socio", "Monto", "Mes Correspondiente", "Plan"]
                header_fill = PatternFill(start_color="0078D7", end_color="0078D7", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=12)
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
                
                # Escribir datos
                for row in range(self.tabla.rowCount()):
                    for col in range(self.tabla.columnCount()):
                        item = self.tabla.item(row, col)
                        value = item.text() if item else ""
                        
                        # Limpiar el formato de monto para Excel
                        if col == 2 and value:  # Columna de Monto
                            # Remover $ y comas, convertir a número
                            value_clean = value.replace('$', '').replace(',', '')
                            try:
                                value = float(value_clean)
                            except:
                                pass
                        
                        cell = ws.cell(row=row+2, column=col+1, value=value)
                        cell.border = border
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                        
                        # Formato especial para montos
                        if col == 2 and isinstance(value, (int, float)):
                            cell.number_format = '$#,##0.00'
                
                # Ajustar ancho de columnas
                ws.column_dimensions['A'].width = 15  # Fecha
                ws.column_dimensions['B'].width = 30  # Socio
                ws.column_dimensions['C'].width = 15  # Monto
                ws.column_dimensions['D'].width = 20  # Mes Correspondiente
                ws.column_dimensions['E'].width = 20  # Plan
                
                # Guardar archivo
                wb.save(filename)
                
                QMessageBox.information(
                    self,
                    "Exportación Exitosa",
                    f"Reporte exportado exitosamente a:\n{filename}\n\n"
                    f"El archivo Excel tiene formato profesional con:\n"
                    f"✓ Encabezados con color\n"
                    f"✓ Celdas separadas correctamente\n"
                    f"✓ Montos formateados como moneda"
                )
                
            except ImportError:
                # Si no está instalado openpyxl, usar CSV mejorado
                QMessageBox.warning(
                    self,
                    "Librería no instalada",
                    "La librería 'openpyxl' no está instalada.\n\n"
                    "Para instalarla, ejecutá en tu terminal:\n"
                    "pip install openpyxl\n\n"
                    "Se exportará en formato CSV como alternativa."
                )
                self._exportar_csv()
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al exportar: {str(e)}")

    def _exportar_csv(self):
        """Exporta a CSV como alternativa."""
        try:
            import csv
            
            filename = f"reporte_gimnasio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            
            with open(filename, 'w', newline='', encoding='utf-8-sig') as file:
                writer = csv.writer(file, delimiter=';')  # Usar ; como separador
                
                # Escribir encabezados
                headers = ["Fecha", "Socio", "Monto", "Mes Correspondiente", "Plan"]
                writer.writerow(headers)
                
                # Escribir datos
                for row in range(self.tabla.rowCount()):
                    row_data = []
                    for col in range(self.tabla.columnCount()):
                        item = self.tabla.item(row, col)
                        row_data.append(item.text() if item else "")
                    writer.writerow(row_data)
            
            QMessageBox.information(
                self,
                "Exportación CSV",
                f"Reporte exportado a CSV:\n{filename}\n\n"
                f"Para abrirlo correctamente en Excel:\n"
                f"1. Abrí Excel\n"
                f"2. Datos > Desde texto/CSV\n"
                f"3. Seleccioná el archivo\n"
                f"4. Elegí delimitador: Punto y coma (;)"
            )
        except Exception as e:
            raise e

    def exportar_pdf(self):
        """Exporta el reporte a PDF."""
        QMessageBox.information(
            self,
            "Próximamente",
            "La exportación a PDF estará disponible próximamente.\n"
            "Por ahora, podés usar la exportación a Excel."
        )'''
"""
views/reports.py
Vista para generar reportes y estadísticas del gimnasio.
Incluye gráficos, exportación a PDF y Excel.
"""

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QLabel, QHBoxLayout, QPushButton,
    QComboBox, QGroupBox, QTableWidget, QTableWidgetItem, QMessageBox
)
from PySide6.QtCore import Qt
import sqlite3
from datetime import datetime, timedelta
from collections import defaultdict


class ReportsView(QWidget):
    def __init__(self, db_path):
        super().__init__()
        self.db_path = db_path
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()

        # Título
        title = QLabel("📊 Reportes y Estadísticas")
        title.setStyleSheet("font-size: 22px; font-weight: bold; margin: 10px;")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        # Panel de filtros
        filtros_group = QGroupBox("🔍 Seleccionar Período")
        filtros_layout = QHBoxLayout()

        self.combo_periodo = QComboBox()
        self.combo_periodo.addItems([
            "Mes Actual",
            "Mes Anterior",
            "Últimos 3 Meses",
            "Últimos 6 Meses",
            "Año Actual",
            "Año Anterior",
            "Últimos 12 Meses",
            "Todo el Tiempo"
        ])
        self.combo_periodo.currentIndexChanged.connect(self.generar_reporte)

        btn_generar = QPushButton("📈 Generar Reporte")
        btn_generar.setStyleSheet("""
            QPushButton {
                padding: 10px 20px;
                background: #2196F3;
                color: white;
                border-radius: 6px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #1976D2;
            }
        """)
        btn_generar.clicked.connect(self.generar_reporte)

        filtros_layout.addWidget(QLabel("Período:"))
        filtros_layout.addWidget(self.combo_periodo)
        filtros_layout.addWidget(btn_generar)
        filtros_layout.addStretch()
        filtros_group.setLayout(filtros_layout)
        layout.addWidget(filtros_group)

        # Panel de resumen financiero
        resumen_group = QGroupBox("💰 Resumen Financiero")
        resumen_layout = QHBoxLayout()

        self.card_ingresos = self._crear_card("Total Ingresos", "$0", "#4CAF50")
        self.card_socios_pagaron = self._crear_card("Socios que Pagaron", "0", "#2196F3")
        self.card_promedio = self._crear_card("Promedio por Pago", "$0", "#FF9800")

        resumen_layout.addWidget(self.card_ingresos)
        resumen_layout.addWidget(self.card_socios_pagaron)
        resumen_layout.addWidget(self.card_promedio)
        resumen_group.setLayout(resumen_layout)
        layout.addWidget(resumen_group)

        # Panel de estadísticas de socios
        socios_group = QGroupBox("👥 Estadísticas de Socios")
        socios_layout = QHBoxLayout()

        self.card_total_socios = self._crear_card("Total Socios", "0", "#9C27B0")
        self.card_nuevos = self._crear_card("Nuevos en Período", "0", "#00BCD4")
        self.card_activos = self._crear_card("Activos", "0", "#4CAF50")

        socios_layout.addWidget(self.card_total_socios)
        socios_layout.addWidget(self.card_nuevos)
        socios_layout.addWidget(self.card_activos)
        socios_group.setLayout(socios_layout)
        layout.addWidget(socios_group)

        # Tabla de detalles
        detalles_group = QGroupBox("📋 Detalle de Pagos del Período")
        detalles_layout = QVBoxLayout()

        self.tabla = QTableWidget()
        self.tabla.setColumnCount(5)
        self.tabla.setHorizontalHeaderLabels([
            "Fecha", "Socio", "Monto", "Mes Correspondiente", "Plan"
        ])
        self.tabla.horizontalHeader().setStretchLastSection(True)
        self.tabla.setAlternatingRowColors(True)
        self.tabla.setStyleSheet("""
            QTableWidget {
                border: 1px solid #ddd;
                border-radius: 6px;
            }
            QHeaderView::section {
                background-color: #0078d7;
                color: white;
                padding: 8px;
                font-weight: bold;
            }
        """)

        detalles_layout.addWidget(self.tabla)
        detalles_group.setLayout(detalles_layout)
        layout.addWidget(detalles_group)

        # Botones de exportación
        export_layout = QHBoxLayout()

        self.btn_export_excel = QPushButton("📊 Exportar a Excel")
        self.btn_export_excel.setStyleSheet("""
            QPushButton {
                padding: 10px 20px;
                background: #4CAF50;
                color: white;
                border-radius: 6px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #388E3C;
            }
        """)
        self.btn_export_excel.clicked.connect(self.exportar_excel)

        self.btn_export_pdf = QPushButton("📄 Exportar a PDF")
        self.btn_export_pdf.setStyleSheet("""
            QPushButton {
                padding: 10px 20px;
                background: #F44336;
                color: white;
                border-radius: 6px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #D32F2F;
            }
        """)
        self.btn_export_pdf.clicked.connect(self.exportar_pdf)

        export_layout.addStretch()
        export_layout.addWidget(self.btn_export_excel)
        export_layout.addWidget(self.btn_export_pdf)
        layout.addLayout(export_layout)

        self.setLayout(layout)
        
        # Generar reporte inicial
        self.generar_reporte()

    def _crear_card(self, titulo, valor, color):
        """Crea una tarjeta de estadística."""
        from PySide6.QtWidgets import QFrame
        
        frame = QFrame()
        frame.setStyleSheet(f"""
            QFrame {{
                background: white;
                border-left: 5px solid {color};
                border-radius: 8px;
                padding: 20px;
            }}
        """)
        
        frame_layout = QVBoxLayout(frame)
        
        lbl_titulo = QLabel(titulo)
        lbl_titulo.setStyleSheet(f"font-size: 14px; color: #666; font-weight: 500;")
        
        lbl_valor = QLabel(valor)
        lbl_valor.setStyleSheet(f"font-size: 32px; font-weight: bold; color: {color};")
        lbl_valor.setAlignment(Qt.AlignCenter)
        
        frame_layout.addWidget(lbl_titulo)
        frame_layout.addWidget(lbl_valor)
        
        frame.lbl_valor = lbl_valor
        return frame

    def _obtener_rango_fechas(self):
        """Obtiene el rango de fechas según el período seleccionado."""
        hoy = datetime.now()
        periodo = self.combo_periodo.currentText()

        if periodo == "Mes Actual":
            inicio = hoy.replace(day=1).strftime('%Y-%m-%d')
            fin = hoy.strftime('%Y-%m-%d')
        elif periodo == "Mes Anterior":
            primer_dia_mes = hoy.replace(day=1)
            ultimo_mes = primer_dia_mes - timedelta(days=1)
            inicio = ultimo_mes.replace(day=1).strftime('%Y-%m-%d')
            fin = ultimo_mes.strftime('%Y-%m-%d')
        elif periodo == "Últimos 3 Meses":
            inicio = (hoy - timedelta(days=90)).strftime('%Y-%m-%d')
            fin = hoy.strftime('%Y-%m-%d')
        elif periodo == "Últimos 6 Meses":
            inicio = (hoy - timedelta(days=180)).strftime('%Y-%m-%d')
            fin = hoy.strftime('%Y-%m-%d')
        elif periodo == "Año Actual":
            inicio = hoy.replace(month=1, day=1).strftime('%Y-%m-%d')
            fin = hoy.strftime('%Y-%m-%d')
        elif periodo == "Año Anterior":
            inicio = hoy.replace(year=hoy.year-1, month=1, day=1).strftime('%Y-%m-%d')
            fin = hoy.replace(year=hoy.year-1, month=12, day=31).strftime('%Y-%m-%d')
        elif periodo == "Últimos 12 Meses":
            inicio = (hoy - timedelta(days=365)).strftime('%Y-%m-%d')
            fin = hoy.strftime('%Y-%m-%d')
        else:  # Todo el Tiempo
            inicio = "2000-01-01"
            fin = hoy.strftime('%Y-%m-%d')

        return inicio, fin

    def generar_reporte(self):
        """Genera el reporte con los datos del período seleccionado."""
        try:
            inicio, fin = self._obtener_rango_fechas()
            
            conn = sqlite3.connect(self.db_path)
            c = conn.cursor()

            # Ingresos totales del período
            c.execute("""
                SELECT SUM(monto), COUNT(*), COUNT(DISTINCT socio_id)
                FROM pagos
                WHERE fecha_pago BETWEEN ? AND ?
            """, (inicio, fin))
            
            resultado = c.fetchone()
            total_ingresos = resultado[0] if resultado[0] else 0
            cantidad_pagos = resultado[1]
            socios_pagaron = resultado[2]

            # Promedio por pago
            promedio = total_ingresos / cantidad_pagos if cantidad_pagos > 0 else 0

            # Actualizar cards financieros
            self.card_ingresos.lbl_valor.setText(f"${total_ingresos:,.2f}")
            self.card_socios_pagaron.lbl_valor.setText(str(socios_pagaron))
            self.card_promedio.lbl_valor.setText(f"${promedio:,.2f}")

            # Estadísticas de socios
            c.execute("SELECT COUNT(*) FROM socios")
            total_socios = c.fetchone()[0]

            # CORRECCIÓN: Usar fecha_inscripcion en lugar de fecha_alta
            c.execute("""
                SELECT COUNT(*) FROM socios
                WHERE fecha_inscripcion BETWEEN ? AND ?
            """, (inicio, fin))
            nuevos_socios = c.fetchone()[0]

            # Socios activos (con al menos un pago en el período)
            c.execute("""
                SELECT COUNT(DISTINCT socio_id) FROM pagos
                WHERE fecha_pago BETWEEN ? AND ?
            """, (inicio, fin))
            socios_activos = c.fetchone()[0]

            # Actualizar cards de socios
            self.card_total_socios.lbl_valor.setText(str(total_socios))
            self.card_nuevos.lbl_valor.setText(str(nuevos_socios))
            self.card_activos.lbl_valor.setText(str(socios_activos))

            # Cargar detalle de pagos
            c.execute("""
                SELECT p.fecha_pago, s.nombre || ' ' || s.apellido,
                       p.monto, p.mes_correspondiente, pl.nombre
                FROM pagos p
                JOIN socios s ON p.socio_id = s.id
                LEFT JOIN planes pl ON s.plan_id = pl.id
                WHERE p.fecha_pago BETWEEN ? AND ?
                ORDER BY p.fecha_pago DESC
            """, (inicio, fin))

            pagos = c.fetchall()
            self.tabla.setRowCount(len(pagos))

            for row, pago in enumerate(pagos):
                self.tabla.setItem(row, 0, QTableWidgetItem(pago[0]))
                self.tabla.setItem(row, 1, QTableWidgetItem(pago[1]))
                self.tabla.setItem(row, 2, QTableWidgetItem(f"${pago[2]:,.2f}"))
                self.tabla.setItem(row, 3, QTableWidgetItem(pago[3]))
                self.tabla.setItem(row, 4, QTableWidgetItem(pago[4] if pago[4] else "Sin plan"))

            conn.close()

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al generar reporte: {str(e)}")

    def exportar_excel(self):
        """Exporta el reporte a Excel con formato."""
        try:
            # Intentar usar openpyxl para Excel real
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                filename = f"reporte_gimnasio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                
                # Crear libro de Excel
                wb = Workbook()
                ws = wb.active
                ws.title = "Reporte de Pagos"
                
                # Escribir encabezados con formato
                headers = ["Fecha", "Socio", "Monto", "Mes Correspondiente", "Plan"]
                header_fill = PatternFill(start_color="0078D7", end_color="0078D7", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=12)
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
                
                # Escribir datos
                for row in range(self.tabla.rowCount()):
                    for col in range(self.tabla.columnCount()):
                        item = self.tabla.item(row, col)
                        value = item.text() if item else ""
                        
                        # Limpiar el formato de monto para Excel
                        if col == 2 and value:  # Columna de Monto
                            # Remover $ y comas, convertir a número
                            value_clean = value.replace('$', '').replace(',', '')
                            try:
                                value = float(value_clean)
                            except:
                                pass
                        
                        cell = ws.cell(row=row+2, column=col+1, value=value)
                        cell.border = border
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                        
                        # Formato especial para montos
                        if col == 2 and isinstance(value, (int, float)):
                            cell.number_format = '$#,##0.00'
                
                # Ajustar ancho de columnas
                ws.column_dimensions['A'].width = 15  # Fecha
                ws.column_dimensions['B'].width = 30  # Socio
                ws.column_dimensions['C'].width = 15  # Monto
                ws.column_dimensions['D'].width = 20  # Mes Correspondiente
                ws.column_dimensions['E'].width = 20  # Plan
                
                # Guardar archivo
                wb.save(filename)
                
                QMessageBox.information(
                    self,
                    "Exportación Exitosa",
                    f"Reporte exportado exitosamente a:\n{filename}\n\n"
                    f"El archivo Excel tiene formato profesional con:\n"
                    f"✓ Encabezados con color\n"
                    f"✓ Celdas separadas correctamente\n"
                    f"✓ Montos formateados como moneda"
                )
                
            except ImportError:
                # Si no está instalado openpyxl, usar CSV mejorado
                QMessageBox.warning(
                    self,
                    "Librería no instalada",
                    "La librería 'openpyxl' no está instalada.\n\n"
                    "Para instalarla, ejecutá en tu terminal:\n"
                    "pip install openpyxl\n\n"
                    "Se exportará en formato CSV como alternativa."
                )
                self._exportar_csv()
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al exportar: {str(e)}")

    def _exportar_csv(self):
        """Exporta a CSV como alternativa."""
        try:
            import csv
            
            filename = f"reporte_gimnasio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            
            with open(filename, 'w', newline='', encoding='utf-8-sig') as file:
                writer = csv.writer(file, delimiter=';')  # Usar ; como separador
                
                # Escribir encabezados
                headers = ["Fecha", "Socio", "Monto", "Mes Correspondiente", "Plan"]
                writer.writerow(headers)
                
                # Escribir datos
                for row in range(self.tabla.rowCount()):
                    row_data = []
                    for col in range(self.tabla.columnCount()):
                        item = self.tabla.item(row, col)
                        row_data.append(item.text() if item else "")
                    writer.writerow(row_data)
            
            QMessageBox.information(
                self,
                "Exportación CSV",
                f"Reporte exportado a CSV:\n{filename}\n\n"
                f"Para abrirlo correctamente en Excel:\n"
                f"1. Abrí Excel\n"
                f"2. Datos > Desde texto/CSV\n"
                f"3. Seleccioná el archivo\n"
                f"4. Elegí delimitador: Punto y coma (;)"
            )
        except Exception as e:
            raise e

    def exportar_pdf(self):
        """Exporta el reporte a PDF con formato profesional."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
            
            filename = f"reporte_gimnasio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            
            # Crear documento PDF
            doc = SimpleDocTemplate(filename, pagesize=A4,
                                   rightMargin=30, leftMargin=30,
                                   topMargin=30, bottomMargin=30)
            
            # Container para los elementos del PDF
            elements = []
            styles = getSampleStyleSheet()
            
            # Estilo personalizado para el título
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#0078d7'),
                spaceAfter=30,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )
            
            # Estilo para subtítulos
            subtitle_style = ParagraphStyle(
                'CustomSubtitle',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=colors.HexColor('#333333'),
                spaceAfter=12,
                spaceBefore=12,
                fontName='Helvetica-Bold'
            )
            
            # Título principal
            titulo = Paragraph("📊 Reporte de Pagos - Gimnasio", title_style)
            elements.append(titulo)
            
            # Información del período
            periodo_text = f"<b>Período:</b> {self.combo_periodo.currentText()}"
            fecha_text = f"<b>Generado el:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            
            info_style = ParagraphStyle(
                'Info',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.HexColor('#666666'),
                alignment=TA_CENTER
            )
            
            elements.append(Paragraph(periodo_text, info_style))
            elements.append(Paragraph(fecha_text, info_style))
            elements.append(Spacer(1, 20))
            
            # Resumen Financiero
            elements.append(Paragraph("💰 Resumen Financiero", subtitle_style))
            
            resumen_data = [
                ['Concepto', 'Valor'],
                ['Total Ingresos', self.card_ingresos.lbl_valor.text()],
                ['Socios que Pagaron', self.card_socios_pagaron.lbl_valor.text()],
                ['Promedio por Pago', self.card_promedio.lbl_valor.text()]
            ]
            
            resumen_table = Table(resumen_data, colWidths=[3*inch, 2*inch])
            resumen_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0078d7')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')])
            ]))
            
            elements.append(resumen_table)
            elements.append(Spacer(1, 20))
            
            # Estadísticas de Socios
            elements.append(Paragraph("👥 Estadísticas de Socios", subtitle_style))
            
            socios_data = [
                ['Concepto', 'Cantidad'],
                ['Total Socios', self.card_total_socios.lbl_valor.text()],
                ['Nuevos en Período', self.card_nuevos.lbl_valor.text()],
                ['Socios Activos', self.card_activos.lbl_valor.text()]
            ]
            
            socios_table = Table(socios_data, colWidths=[3*inch, 2*inch])
            socios_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0078d7')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')])
            ]))
            
            elements.append(socios_table)
            elements.append(Spacer(1, 20))
            
            # Detalle de Pagos
            elements.append(Paragraph("📋 Detalle de Pagos del Período", subtitle_style))
            
            # Preparar datos de la tabla
            data = [['Fecha', 'Socio', 'Monto', 'Mes Corresp.', 'Plan']]
            
            for row in range(self.tabla.rowCount()):
                row_data = []
                for col in range(self.tabla.columnCount()):
                    item = self.tabla.item(row, col)
                    text = item.text() if item else ""
                    # Acortar nombres muy largos para que quepan
                    if col == 1 and len(text) > 25:
                        text = text[:22] + "..."
                    if col == 4 and len(text) > 15:
                        text = text[:12] + "..."
                    row_data.append(text)
                data.append(row_data)
            
            # Si hay muchos registros, limitar para que quepa en el PDF
            max_rows = 30
            if len(data) > max_rows + 1:
                data = data[:max_rows + 1]
                elements.append(Paragraph(
                    f"<i>Nota: Se muestran los primeros {max_rows} registros de {self.tabla.rowCount()} totales.</i>",
                    info_style
                ))
                elements.append(Spacer(1, 10))
            
            # Crear tabla con anchos de columna ajustados
            detail_table = Table(data, colWidths=[1.2*inch, 1.8*inch, 1*inch, 1.2*inch, 1.3*inch])
            detail_table.setStyle(TableStyle([
                # Encabezado
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0078d7')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                
                # Datos
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Fecha centrada
                ('ALIGN', (1, 1), (1, -1), 'LEFT'),    # Nombre a la izquierda
                ('ALIGN', (2, 1), (2, -1), 'RIGHT'),   # Monto a la derecha
                ('ALIGN', (3, 1), (3, -1), 'CENTER'),  # Mes centrado
                ('ALIGN', (4, 1), (4, -1), 'LEFT'),    # Plan a la izquierda
                
                # Bordes y colores alternados
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')]),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ]))
            
            elements.append(detail_table)
            
            # Pie de página
            elements.append(Spacer(1, 30))
            footer_style = ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.HexColor('#999999'),
                alignment=TA_CENTER
            )
            footer_text = "Sistema de Gestión de Gimnasio | Generado automáticamente"
            elements.append(Paragraph(footer_text, footer_style))
            
            # Construir PDF
            doc.build(elements)
            
            QMessageBox.information(
                self,
                "✅ Exportación Exitosa",
                f"Reporte PDF generado exitosamente:\n{filename}\n\n"
                f"El archivo incluye:\n"
                f"✓ Resumen financiero completo\n"
                f"✓ Estadísticas de socios\n"
                f"✓ Detalle de todos los pagos\n"
                f"✓ Formato profesional con colores"
            )
            
        except ImportError:
            QMessageBox.warning(
                self,
                "Librería no instalada",
                "La librería 'reportlab' no está instalada.\n\n"
                "Para instalarla, ejecutá en tu terminal:\n"
                "pip install reportlab\n\n"
                "Luego podrás exportar reportes en formato PDF."
            )
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al exportar PDF: {str(e)}")