# views/payments_complete_view.py
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QLabel, QTableWidget, QTableWidgetItem,
    QHBoxLayout, QPushButton, QComboBox, QLineEdit, QMessageBox,
    QGroupBox, QDoubleSpinBox, QFrame, QScrollArea
)
from PySide6.QtCore import Qt, QTimer
from PySide6.QtGui import QColor
from datetime import datetime
from .editar_pago_dialog import EditarPagoDialog
from .historial_socio_dialog import HistorialSocioDialog


class PaymentsCompleteView(QWidget):
    """Vista completa de pagos de socios."""
    def __init__(self, controller, email_service=None):
        super().__init__()
        self.controller = controller
        self.email_service = email_service
        self.pago_seleccionado = None
        self.init_ui()

        # Autoactualizar cada 60 segundos
        self.timer = QTimer()
        self.timer.timeout.connect(self.actualizar_datos)
        self.timer.start(60000)

    def init_ui(self):
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)

        main_widget = QWidget()
        layout = QVBoxLayout(main_widget)
        layout.setSpacing(20)

        # --- Título ---
        title = QLabel("💰 Gestión de Pagos de Socios")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("""
            font-size: 24px;
            font-weight: bold;
            color: #0078d7;
            padding: 15px;
        """)
        layout.addWidget(title)

        # --- Resumen del mes ---
        stats_group = QGroupBox("📊 Resumen del Mes Actual")
        stats_layout = QHBoxLayout()

        mes_inicio = datetime.now().replace(day=1).strftime('%Y-%m-%d')
        mes_fin = datetime.now().strftime('%Y-%m-%d')
        stats = self.controller.obtener_estadisticas(mes_inicio, mes_fin)

        self.card_total_mes = self._crear_card("Total del Mes", f"${stats['total_ingresos']:,.0f}", "#4CAF50")
        self.card_pagos_mes = self._crear_card("Pagos Registrados", str(stats['total_pagos']), "#2196F3")
        self.card_socios_mes = self._crear_card("Socios que Pagaron", str(stats['socios_pagaron']), "#FF9800")
        self.card_promedio = self._crear_card("Promedio", f"${stats['promedio_pago']:,.0f}", "#9C27B0")

        for card in [self.card_total_mes, self.card_pagos_mes, self.card_socios_mes, self.card_promedio]:
            stats_layout.addWidget(card)

        stats_group.setLayout(stats_layout)
        layout.addWidget(stats_group)

        # --- Formulario rápido ---
        form_group = QGroupBox("➕ Registrar Nuevo Pago")
        form_layout = QHBoxLayout()
        self.cmb_socio = QComboBox()
        self.cargar_socios()

        self.spin_monto = QDoubleSpinBox()
        self.spin_monto.setRange(0, 1000000)
        self.spin_monto.setPrefix("$ ")
        self.spin_monto.setValue(5000)
        self.spin_monto.setSingleStep(100)

        self.cmb_mes = QComboBox()
        self.cargar_meses()

        self.cmb_metodo = QComboBox()
        self.cmb_metodo.addItems(["Efectivo", "Transferencia", "Débito", "Crédito", "MercadoPago", "Otro"])

        btn_registrar = QPushButton("✅ Registrar Pago")
        btn_registrar.setStyleSheet("""
            QPushButton {
                padding: 12px 24px;
                background: #28a745;
                color: white;
                border-radius: 8px;
                font-weight: bold;
            }
            QPushButton:hover { background: #218838; }
        """)
        btn_registrar.clicked.connect(self.registrar_pago_rapido)

        form_layout.addWidget(QLabel("Socio:"))
        form_layout.addWidget(self.cmb_socio)
        form_layout.addWidget(QLabel("Monto:"))
        form_layout.addWidget(self.spin_monto)
        form_layout.addWidget(QLabel("Mes:"))
        form_layout.addWidget(self.cmb_mes)
        form_layout.addWidget(QLabel("Método:"))
        form_layout.addWidget(self.cmb_metodo)
        form_layout.addWidget(btn_registrar)
        form_group.setLayout(form_layout)
        layout.addWidget(form_group)

        # --- Filtros ---
        filtros_group = QGroupBox("🔍 Filtros de Búsqueda")
        filtros_layout = QHBoxLayout()

        self.cmb_filtro_socio = QComboBox()
        self.cmb_filtro_socio.addItem("Todos los socios", None)
        self.cargar_socios_filtro()

        self.cmb_filtro_mes = QComboBox()
        self.cmb_filtro_mes.addItem("Todos los meses", "")
        for i, mes in enumerate(["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                                 "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"], 1):
            self.cmb_filtro_mes.addItem(mes, str(i).zfill(2))

        self.cmb_filtro_anio = QComboBox()
        self.cmb_filtro_anio.addItem("Todos los años", "")
        anio_actual = datetime.now().year
        for anio in range(anio_actual - 5, anio_actual + 1):
            self.cmb_filtro_anio.addItem(str(anio), str(anio))
        self.cmb_filtro_anio.setCurrentText(str(anio_actual))

        btn_buscar = QPushButton("🔎 Buscar")
        btn_buscar.clicked.connect(self.aplicar_filtros)
        btn_limpiar = QPushButton("🔄 Limpiar")
        btn_limpiar.clicked.connect(self.limpiar_filtros)

        for w in [
            QLabel("Socio:"), self.cmb_filtro_socio,
            QLabel("Mes:"), self.cmb_filtro_mes,
            QLabel("Año:"), self.cmb_filtro_anio,
            btn_buscar, btn_limpiar
        ]:
            filtros_layout.addWidget(w)
        filtros_layout.addStretch()

        filtros_group.setLayout(filtros_layout)
        layout.addWidget(filtros_group)

        # --- Tabla ---
        tabla_group = QGroupBox("📋 Historial de Pagos")
        tabla_layout = QVBoxLayout()
        self.tabla = QTableWidget()
        self.tabla.setColumnCount(8)
        self.tabla.setHorizontalHeaderLabels([
            "ID", "Socio", "Monto", "Fecha de Pago",
            "Mes Correspondiente", "Método", "Observaciones", "Acciones"
        ])
        self.tabla.setAlternatingRowColors(True)
        self.tabla.setSelectionBehavior(QTableWidget.SelectRows)
        self.tabla.horizontalHeader().setStretchLastSection(True)
        self.tabla.setMinimumHeight(300)
        tabla_layout.addWidget(self.tabla)
        tabla_group.setLayout(tabla_layout)
        layout.addWidget(tabla_group)

        # --- Botones de acción ---
        btn_layout = QHBoxLayout()
        self.btn_editar = QPushButton("✏️ Editar Pago")
        self.btn_eliminar = QPushButton("🗑️ Eliminar Pago")
        self.btn_historial = QPushButton("📜 Ver Historial")
        self.btn_exportar = QPushButton("📊 Exportar CSV")

        self.btn_editar.clicked.connect(self.editar_pago)
        self.btn_eliminar.clicked.connect(self.eliminar_pago)
        self.btn_historial.clicked.connect(self.ver_historial_socio)
        self.btn_exportar.clicked.connect(self.exportar_excel)

        for b in [self.btn_editar, self.btn_eliminar, self.btn_historial, self.btn_exportar]:
            btn_layout.addWidget(b)
        layout.addLayout(btn_layout)

        scroll.setWidget(main_widget)
        main_layout = QVBoxLayout(self)
        main_layout.addWidget(scroll)
        self.cargar_pagos()

    # --- Métodos auxiliares ---
    def _crear_card(self, titulo, valor, color):
        frame = QFrame()
        frame.setStyleSheet(f"QFrame{{border-top:5px solid {color};border-radius:12px;padding:20px;}}")
        lay = QVBoxLayout(frame)
        lbl_t = QLabel(titulo)
        lbl_t.setAlignment(Qt.AlignCenter)
        lbl_v = QLabel(valor)
        lbl_v.setAlignment(Qt.AlignCenter)
        lbl_v.setStyleSheet(f"font-size:26px;font-weight:bold;color:{color};")
        lay.addWidget(lbl_t)
        lay.addWidget(lbl_v)
        frame.lbl_valor = lbl_v
        return frame

    def cargar_socios(self):
        self.cmb_socio.clear()
        for s in self.controller.obtener_socios():
            self.cmb_socio.addItem(f"{s[1]} {s[2]}", s[0])

    def cargar_socios_filtro(self):
        for s in self.controller.obtener_socios():
            self.cmb_filtro_socio.addItem(f"{s[1]} {s[2]}", s[0])

    def cargar_meses(self):
        meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                 "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        anio = datetime.now().year
        for i, mes in enumerate(meses, 1):
            self.cmb_mes.addItem(f"{mes} {anio}", f"{anio}-{str(i).zfill(2)}")
        self.cmb_mes.setCurrentIndex(datetime.now().month - 1)

    def cargar_pagos(self):
        self._mostrar_pagos(self.controller.obtener_todos_los_pagos())

    def _mostrar_pagos(self, pagos):
        self.tabla.setRowCount(len(pagos))
        for r, p in enumerate(pagos):
            self.tabla.setItem(r, 0, QTableWidgetItem(str(p[0])))
            self.tabla.setItem(r, 1, QTableWidgetItem(p[1]))
            monto = QTableWidgetItem(f"${p[2]:,.2f}")
            monto.setForeground(QColor(76, 175, 80))
            self.tabla.setItem(r, 2, monto)
            for i, v in enumerate(p[3:7], 3):
                self.tabla.setItem(r, i, QTableWidgetItem(str(v or "-")))

    def registrar_pago_rapido(self):
        socio_id = self.cmb_socio.currentData()
        monto = self.spin_monto.value()
        mes = self.cmb_mes.currentData()
        metodo = self.cmb_metodo.currentText().lower()
        if not socio_id or monto <= 0:
            QMessageBox.warning(self, "Error", "Seleccioná socio y monto válido")
            return
        if self.controller.verificar_duplicado(socio_id, mes):
            if QMessageBox.question(self, "Pago duplicado", "Ya existe un pago para este mes. ¿Registrar igual?",
                                    QMessageBox.Yes | QMessageBox.No) == QMessageBox.No:
                return
        self.controller.registrar_pago(socio_id, monto, mes, metodo)
        self.cargar_pagos()
        QMessageBox.information(self, "Éxito", "Pago registrado correctamente")

    def editar_pago(self):
        fila = self.tabla.currentRow()
        if fila < 0: return QMessageBox.warning(self, "Error", "Seleccioná un pago")
        pago_id = int(self.tabla.item(fila, 0).text())
        pago = self.controller.obtener_pago(pago_id)
        if not pago: return
        dialogo = EditarPagoDialog(pago, self)
        if dialogo.exec():
            monto, mes, metodo, obs = dialogo.obtener_datos()
            self.controller.actualizar_pago(pago_id, monto, mes, metodo, obs)
            self.cargar_pagos()

    def eliminar_pago(self):
        fila = self.tabla.currentRow()
        if fila < 0: return QMessageBox.warning(self, "Error", "Seleccioná un pago")
        pago_id = int(self.tabla.item(fila, 0).text())
        if QMessageBox.question(self, "Eliminar", "¿Eliminar este pago?", QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            self.controller.eliminar_pago(pago_id)
            self.cargar_pagos()

    def ver_historial_socio(self):
        fila = self.tabla.currentRow()
        if fila < 0: return QMessageBox.warning(self, "Error", "Seleccioná un pago")
        socio_id = self.controller.obtener_todos_los_pagos()[fila][7]
        socio_nombre = self.tabla.item(fila, 1).text()
        HistorialSocioDialog(self.controller, socio_id, socio_nombre, self).exec()

    def exportar_excel(self):
        """Exporta los pagos a un archivo Excel con formato, en carpeta Excels."""
        try:
            import os
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from datetime import datetime

            # Crear carpeta si no existe
            carpeta = "Excels"
            os.makedirs(carpeta, exist_ok=True)

            # Obtener nombre descriptivo
            mes_idx = self.cmb_filtro_mes.currentIndex()
            mes_nombre = self.cmb_filtro_mes.currentText()
            anio = self.cmb_filtro_anio.currentText() or str(datetime.now().year)

            # Si no se filtró un mes específico, usar el mes actual
            if mes_idx == 0 or mes_nombre == "Todos los meses":
                mes_nombre = datetime.now().strftime("%B").capitalize()
            nombre_archivo = f"Pago {mes_nombre} {anio}.xlsx"
            ruta_archivo = os.path.join(carpeta, nombre_archivo)

            # Crear libro y hoja
            wb = Workbook()
            ws = wb.active
            ws.title = "Pagos"

            # Encabezados
            headers = [
                self.tabla.horizontalHeaderItem(i).text()
                for i in range(self.tabla.columnCount() - 1)
            ]
            ws.append(headers)

            # Estilo del encabezado
            header_fill = PatternFill(start_color="0078D7", end_color="0078D7", fill_type="solid")
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = header_fill
            ws.row_dimensions[1].height = 25

            # Agregar datos
            for r in range(self.tabla.rowCount()):
                row_data = []
                for c in range(self.tabla.columnCount() - 1):
                    item = self.tabla.item(r, c)
                    row_data.append(item.text() if item else "")
                ws.append(row_data)

            # Ajustar ancho de columnas
            for col in ws.columns:
                max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[col[0].column_letter].width = max_len + 4

            # Guardar archivo
            wb.save(ruta_archivo)
            QMessageBox.information(self, "✅ Exportado correctamente",
                                    f"Archivo creado en:\n{os.path.abspath(ruta_archivo)}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudo exportar a Excel:\n{e}")



    def aplicar_filtros(self):
        socio = self.cmb_filtro_socio.currentData()
        mes = self.cmb_filtro_mes.currentData()
        anio = self.cmb_filtro_anio.currentData()
        pagos = self.controller.buscar_pagos(socio, None, None, mes, anio)
        self._mostrar_pagos(pagos)

    def limpiar_filtros(self):
        self.cmb_filtro_socio.setCurrentIndex(0)
        self.cmb_filtro_mes.setCurrentIndex(0)
        self.cmb_filtro_anio.setCurrentIndex(0)
        self.cargar_pagos()

    def actualizar_datos(self):
        self.cargar_pagos()
